import pandas as pd
from googletrans import Translator
import openpyxl as op



#Import fichier excel source
#Import fichier excel source
df_source=pd.read_excel("C:/Users/adel-mounir.achir/Downloads/GDSN_and_Shared_Code_Lists_3p1p23_v2_December_19_2022.xlsx",sheet_name="CodeList")
#Filter df_source
filter_column="Code List"
filter_value="PackagingMaterialTypeCode"
filt = (df_source[filter_column] == filter_value)
filtered_df=(df_source[filt].iloc[1: , :]) #enlever la première ligne
filtered_df = filtered_df.reset_index(drop=True)

print(type(filtered_df))
#Get Rows And Columns
rows = len(filtered_df.axes[0])+1
cols = len(filtered_df.axes[1])

# Print the number of rows and columns
print("Number of Rows: " + str(rows))
print("Number of Columns: " + str(cols))

df=pd.read_excel("C:/Users/adel-mounir.achir/Downloads/Groupe de standardisation AGEC/Packaging Codification Guide FR ENG.xlsx",sheet_name="PackageTypeCode")



#Traduction et Insertion dans le fichier excel de base 
file_name="C:/Users/adel-mounir.achir/Downloads/Groupe de standardisation AGEC/Packaging Codification Guide FR ENG.xlsx"
wb = op.load_workbook(filename = file_name)
ws = wb.worksheets[2]
print(ws)
detector = Translator()



for i in range(3,rows+1):
    #first : codevalue
    ws.cell(row=i,column=1).value=filtered_df.loc[i-3,"Code Value"]
    #second : code name
    ws.cell(row=i,column=2).value=filtered_df.loc[i-3,"Code Name"]
    #Traduction code name
    ws.cell(row=i,column=3).value=detector.translate(filtered_df.loc[i-3,"Code Name"],dest="fr").text
    #third : Definition
    ws.cell(row=i,column=5).value=filtered_df.loc[i-3,"Definition"]
    #Traduction Definition
    ws.cell(row=i,column=6).value=detector.translate(filtered_df.loc[i-3,"Definition"],dest="fr").text
    
wb.save(filename=file_name)