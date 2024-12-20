import numpy as np
import pandas as pd
import openpyxl as op
from openpyxl import load_workbook, Workbook
import copy
import time
from openpyxl.styles import Font, Alignment, PatternFill
start_time = time.time()
################################################################
################################################################
#############Paramétrage########################################
################################################################
################################################################
#Paramètres du codes
chemin="C:/Users/adel-mounir.achir/OneDrive - GS1/Bureau/Mes documents/Codes and automations/Fiche produit PGC Relooking/"
chemain_Export=chemin+"Exports/"
Nom_Fp_Principale="FicheProduit_3.1.28_V2.xlsx"
sheet_name="Fiche-Produit PGC 3.1.28"
version_number="3.1.28"

Static_Sheets_names=[ 'MENTIONS DE DANGER ', 'CONSEILS DE PRUDENCE', 'Labels_Accreditations_FRANCE', 'Nouveaux_Labels_Accreditations', 'Niveaux logistiques', 'Reco Libellé long', 'Min et multiple Cde', 'Formats promo speciaux', 'Gestion attributs agiles', 
                     'Gestion des composants sans GTI']
Dynamic_Sheets_name=['Sommaire', 'Lisez-moi', 'Mapping GPC vs Catégorie GDM', 'Codes utilisables en France', 'Ecarts PGC FR '+version_number, 'Autres codes spécifiques France', 'DONNEES SUPPRIMEES', 'Liste des attributs Agiles FR']
Fiche_Produit_Tous=['Sommaire', 'Lisez-moi', 'Mapping GPC vs Catégorie GDM','Fiche-Produit PGC '+version_number, 'Codes utilisables en France', 'Ecarts PGC FR '+version_number, 'Autres codes spécifiques France', 'DONNEES SUPPRIMEES', 'Liste des attributs Agiles FR']

wb_source = op.load_workbook(chemin+"Templates/"+Nom_Fp_Principale)
df=pd.read_excel(chemin+"Templates/"+Nom_Fp_Principale,sheet_name=sheet_name,header=2)


output_file_annexes = chemain_Export+'Annexes.xlsx'
output_file_fp_AH=chemain_Export+'FicheProduit_'+version_number+'_PROFIL_Alimentation_Humaine.xlsx'
output_file_fp_DPH=chemain_Export+'FicheProduit_'+version_number+'_PROFIL_DPH.xlsx'
output_file_fp_AA=chemain_Export+'FicheProduit_'+version_number+'_PROFIL_Alimentation_Animale.xlsx'
output_file_fp_tabac=chemain_Export+'FicheProduit_'+version_number+'_PROFIL_TABAC.xlsx'
output_file_fp_BA=chemain_Export+'FicheProduit_'+version_number+'_PROFIL_Boissons_Alcolisées.xlsx'
output_file_fp_VS=chemain_Export+'FicheProduit_'+version_number+'_PROFIL_Vins et Spiritueux.xlsx'
output_file_fp_COS=chemain_Export+'FicheProduit_'+version_number+'_PROFIL_Cosmétiques.xlsx'
output_file_fp_PGC=chemain_Export+'FicheProduit_'+version_number+'_PROFIL_PGC.xlsx'


print("First Step :Variables declared")

################################################################
################################################################
###################Functions####################################
################################################################
################################################################

#Function to copy a sheet
# Function to copy a whole sheet, including formatting and structure
def copy_sheet(source_sheet, target_sheet):
    # Copy all cells, including values and formatting
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = target_sheet[cell.coordinate]
            new_cell.value = cell.value
            
            # Copy formatting
            if cell.has_style:
                new_cell.font = copy.copy(cell.font)
                new_cell.border = copy.copy(cell.border)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy.copy(cell.protection)
                new_cell.alignment = copy.copy(cell.alignment)
    # Copy merged cells
    for merged_cell_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_cell_range))

    # Copy freeze panes
    target_sheet.freeze_panes = source_sheet.freeze_panes

    # Copy row heights
    for row_num, row_dim in source_sheet.row_dimensions.items():
        target_sheet.row_dimensions[row_num].height = row_dim.height

    # Copy column widths
    for col_letter, col_dim in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[col_letter].width = col_dim.width
        
    
    # # Copy sheet properties (dimensions, merged cells, etc.)
    # target_sheet.sheet_properties = copy.copy(source_sheet.sheet_properties)
    # target_sheet.merged_cells = copy.copy(source_sheet.merged_cells)
    # target_sheet.freeze_panes = source_sheet.freeze_panes
    
    # Copy dimensions
    # for dim in source_sheet.row_dimensions:
    #     target_sheet.row_dimensions[dim] = copy.copy(source_sheet.row_dimensions[dim])
    
    # for dim in source_sheet.column_dimensions:
    #     target_sheet.column_dimensions[dim] = copy.copy(source_sheet.column_dimensions[dim])

#Recréation des feuilles communes à toutes les fiches produits
#Méthode 1 Création d'un doc Annexe
def Extract_sheets(wb_source, Static_Sheets_names, output_file):
    # Create new workbook to copy sheets into
    wb_target = Workbook()

    # Remove the default sheet created by Workbook()
    if "Sheet" in wb_target.sheetnames:
        del wb_target["Sheet"]

    # Ensure at least one sheet is added to avoid errors
    added_sheet = False

    # Iterate through all the sheet names in the Static_Sheets_names
    for sheet_name in Static_Sheets_names:
        source_sheet = wb_source[sheet_name]
        
        # Create a new sheet in the target workbook
        wb_target.create_sheet(title=sheet_name)
        target_sheet = wb_target[sheet_name]
        added_sheet = True  # A sheet has been added
        copy_sheet(source_sheet,target_sheet)
        print("Second Step : "+sheet_name+" copied")

    # Save the new workbook with all sheets and formatting
    wb_target.save(output_file)

    return f"All sheets with formatting copied to {output_file}"

# Execute the example

#Function to format an excel data sheet
# Load the workbook and the newly added sheet to apply formatting
def design_data_sheet(output_file,sheet_name):
    wb = op.load_workbook(output_file)
    ws = wb[sheet_name]  # Access the new sheet by name

    # Define your desired formatting
    header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")  # Light blue background
    header_font = Font(name="Verdana",bold=True, color="000000")  # Bold text for header
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Apply formatting to the header row (first row)
    for cell in ws[1]:  # First row (header row)
        cell.fill = header_fill  # Apply background color
        cell.font = header_font  # Apply bold font
        cell.alignment = center_alignment  # Center-align text

        # Apply formatting to data cells (center-align and wrap text)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)  # Left-align and wrap text

    # Adjust column width automatically to fit content (optional step, if necessary)
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter  # Get the column name
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    # Save the workbook with the applied formatting
    wb.save(output_file)


#Création de la fiche produit
def creation_Fiche(fiche_produit_name,output_file,Dynamic_Sheets_name,df):
    if fiche_produit_name=="PGC":
        Extract_sheets(wb_source, Fiche_Produit_Tous, output_file)
        return
    
    Extract_sheets(wb_source, Dynamic_Sheets_name, output_file)
    wb_FP = op.load_workbook(output_file,data_only=True)

    # Filter the DataFrame based on your condition : Alimentation animale
    if fiche_produit_name == "ALIMENTATION_ANIMALE":
        df2 = df[(df['Catégorie(s) de produit(s) concerné(s)'].str.contains("ALIMENTATION_ANIMALE", case=False, na=False)) | 
         (df['Catégorie(s) de produit(s) concerné(s)'] == "TOUS_PRODUITS")]    
        # Use ExcelWriter to write the DataFrame to the workbook
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:    
            df2.to_excel(writer, sheet_name=fiche_produit_name, index=False)
     # Filter the DataFrame based on your condition : Alimentation Humaine
    elif fiche_produit_name == "ALIMENTATION_HUMAINE":
        df2 = df[(df['Catégorie(s) de produit(s) concerné(s)'].str.contains("ALIMENTATION_HUMAINE", case=False, na=False)) | 
         (df['Catégorie(s) de produit(s) concerné(s)'] == "TOUS_PRODUITS")]       
        # Use ExcelWriter to write the DataFrame to the workbook
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:    
            df2.to_excel(writer, sheet_name=fiche_produit_name, index=False)
    # Filter the DataFrame based on your condition : DPH
    elif fiche_produit_name == "DROGUERIE_PARFUMERIE_&HYGIENE":
        df2 = df[(df['Catégorie(s) de produit(s) concerné(s)'].str.contains("DROGUERIE_PARFUMERIE_HYGIENE", case=False, na=False)) | 
         (df['Catégorie(s) de produit(s) concerné(s)'] == "TOUS_PRODUITS")]     
        # Use ExcelWriter to write the DataFrame to the workbook
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:    
            df2.to_excel(writer, sheet_name=fiche_produit_name, index=False)
        # Filter the DataFrame based on your condition : BOISSONS ALCO
    elif fiche_produit_name == "BOISSONS_ALCOOLISEES":
        df2 = df[(df['Catégorie(s) de produit(s) concerné(s)'].str.contains("BOISSONS_ALCOOLISEES", case=False, na=False)) | 
         (df['Catégorie(s) de produit(s) concerné(s)'] == "TOUS_PRODUITS")]      
        # Use ExcelWriter to write the DataFrame to the workbook
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:    
            df2.to_excel(writer, sheet_name=fiche_produit_name, index=False)
        # Filter the DataFrame based on your condition : TABAC
    elif fiche_produit_name == "TABAC":
        df2 = df[(df['Catégorie(s) de produit(s) concerné(s)'].str.contains("TABAC", case=False, na=False)) | 
         (df['Catégorie(s) de produit(s) concerné(s)'] == "TOUS_PRODUITS")]     
         
        # Use ExcelWriter to write the DataFrame to the workbook
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:    
            df2.to_excel(writer, sheet_name=fiche_produit_name, index=False)
        # Filter the DataFrame based on your condition : TABAC
    elif fiche_produit_name == "Vins & Spi B2C":
        df2 = df[(df['Attribut Wine & Spirits consumer unit ?']=="OUI")]     
        # Use ExcelWriter to write the DataFrame to the workbook
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:    
            df2.to_excel(writer, sheet_name=fiche_produit_name, index=False)
    elif fiche_produit_name == "Cosmétiques":
        df2 = df[(df['Catégorie(s) de produit(s) concerné(s)'].str.contains("DROGUERIE_PARFUMERIE_HYGIENE", case=False, na=False)) | 
         (df['Catégorie(s) de produit(s) concerné(s)'] == "TOUS_PRODUITS")]     
        # Use ExcelWriter to write the DataFrame to the workbook
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a') as writer:    
            df2.to_excel(writer, sheet_name=fiche_produit_name, index=False)

    design_data_sheet(output_file,fiche_produit_name)

################################################################
################################################################
#############Création des fiches produits#######################
################################################################
################################################################
# Extract_sheets(wb_source, Static_Sheets_names, output_file_annexes)
# print("Third Step : Annexe created")
            
# creation_Fiche("ALIMENTATION_ANIMALE",output_file_fp_AA,Dynamic_Sheets_name,df)
# print("Final Step : Fiche produit Alimentation animale created")
# creation_Fiche("ALIMENTATION_HUMAINE",output_file_fp_AH,Dynamic_Sheets_name,df)
# print("Final Step : Fiche produit Alimentation Humaine created")
# creation_Fiche("DROGUERIE_PARFUMERIE_&HYGIENE",output_file_fp_DPH,Dynamic_Sheets_name,df)
# print("Final Step : Fiche produit DPH created")
# creation_Fiche("BOISSONS_ALCOOLISEES",output_file_fp_BA,Dynamic_Sheets_name,df)
# print("Final Step : Fiche produit Boissons alcolisées created")
# creation_Fiche("TABAC",output_file_fp_tabac,Dynamic_Sheets_name,df)
# print("Final Step : Fiche produit TABAC created")
# creation_Fiche("Vins & Spi B2C",output_file_fp_VS,Dynamic_Sheets_name,df)
# print("Final Step : Fiche produit Vins & Spi B2C created")
# creation_Fiche("Cosmétiques",output_file_fp_COS,Dynamic_Sheets_name,df)
# print("Final Step : Fiche produit Cosmétiques created")
# creation_Fiche("PGC",output_file_fp_PGC,Dynamic_Sheets_name,df)
# # print("Final Step : Fiche PGC created")
# end_time = time.time()
# execution_time = end_time - start_time
# print(f"Execution time: {execution_time} seconds, {execution_time/60} minutes")

# df=pd.read_excel(chemin+Nom_Fp_Principale,sheet_name=sheet_name,header=2)
# df2 = df[(df['Catégorie(s) de produit(s) concerné(s)'] == "ALIMENTATION_ANIMALE") | 
# (df['Catégorie(s) de produit(s) concerné(s)'] == "TOUS_PRODUITS")]
# print(df2)

# Fonctions pour les boutons
